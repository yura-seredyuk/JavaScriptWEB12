from openpyxl import Workbook


data = {
        "exchangeRate":
        [
            {"baseCurrency":"UAH","currency":"CHF","saleRateNB":15.6389750,"purchaseRateNB":15.6389750,"saleRate":17.0000000,"purchaseRate":15.5000000},
            {"baseCurrency":"UAH","currency":"EUR","saleRateNB":18.7949200,"purchaseRateNB":18.7949200,"saleRate":20.0000000,"purchaseRate":19.2000000},
            {"baseCurrency":"UAH","currency":"GBP","saleRateNB":23.6324910,"purchaseRateNB":23.6324910,"saleRate":25.8000000,"purchaseRate":24.0000000},
            {"baseCurrency":"UAH","currency":"PLZ","saleRateNB":4.4922010,"purchaseRateNB":4.4922010,"saleRate":5.0000000,"purchaseRate":4.2000000},
            {"baseCurrency":"UAH","currency":"USD","saleRateNB":15.0564130,"purchaseRateNB":15.0564130,"saleRate":15.7000000,"purchaseRate":15.3500000},
            {"baseCurrency":"UAH","currency":"CAD","saleRateNB":13.2107400,"purchaseRateNB":13.2107400,"saleRate":15.0000000,"purchaseRate":13.0000000}
        ]
    }
workbook = Workbook()
sheet = workbook.active
sheet.title = 'PB_data'

header_row = 1

for index, item in enumerate(list(data["exchangeRate"][0].keys()), 1):
    sheet.cell(row=header_row, column=index).value = item


for r_ind, row in enumerate(data["exchangeRate"], header_row+1):
    for c_index, item in enumerate(row, 1):
        sheet.cell(row=r_ind, column=c_index).value = row[item]


# sheet["A1"] = 'Data'

workbook.save(filename="files/file.xlsx")
