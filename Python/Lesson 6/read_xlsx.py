from pprint import pprint
from openpyxl import load_workbook


workbook = load_workbook(filename="files/file.xlsx")
workbook.sheetnames

# print(workbook.sheetnames)

sheet = workbook.active
# print(sheet)
# print(sheet["A1"])
# print(sheet["A1"].value)
# print(sheet.cell(row=1, column=1).value)
# print(sheet.cell(1, 1).value)
# print(sheet["A"])
# for item in sheet["C"]:
#     print(item.value)

# for row in sheet.iter_rows(min_row=1, max_row=7,
#                             min_col=1, max_col=6):
#     print(row)


# for value in sheet.iter_rows(min_row=1, max_row=7,
#                             min_col=1, max_col=6, values_only=True):
#     print(value)


data = {}
data["exchangeRate"] = []

for value in sheet.iter_rows(min_row=1, max_row=1,
                            min_col=1, max_col=6, values_only=True):
    keys = value

# print(keys)

for value in sheet.iter_rows(min_row=2, max_row=7,
                            min_col=1, max_col=6, values_only=True):
    # row = {}
    # for key_id, item in enumerate(value):
    #     row.update({keys[key_id]:item})
    # data["exchangeRate"].append(row)
    data["exchangeRate"].append({key:item for (key, item) in zip(keys, value)})
    
pprint(data)